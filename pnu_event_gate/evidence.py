from __future__ import annotations

import csv
import io
import json
import math
import re
import shutil
import subprocess
import tempfile
import unicodedata
import zipfile
from dataclasses import asdict, dataclass
from html.parser import HTMLParser
from pathlib import Path
from typing import Any


DEFAULT_CHUNK_CHARS = 1800
DEFAULT_MAX_EVIDENCE_CHARS = 2_000_000
DEFAULT_MAX_ZIP_ENTRY_BYTES = 20_000_000
DEFAULT_MAX_ZIP_TOTAL_BYTES = 80_000_000
VISUAL_PENDING_TEXT = "[visual page pending transcription]"
MIN_LOCAL_OCR_CHARS = 12

STOP_TERMS = {
    "그때",
    "되면",
    "대한",
    "관련",
    "알려줘",
    "알려주세요",
    "에서",
    "으로",
    "이번",
    "있으면",
    "해당",
}


@dataclass(frozen=True)
class EvidenceChunk:
    id: str
    source_name: str
    text: str
    kind: str = "text"
    page: int | None = None
    row: int | None = None
    source_url: str | None = None
    local_path: str | None = None
    extraction_warning: str | None = None

    def to_json(self) -> dict[str, Any]:
        return {key: value for key, value in asdict(self).items() if value is not None}


@dataclass(frozen=True)
class EvidenceBundle:
    chunks: list[EvidenceChunk]
    warnings: list[str]
    redaction_counts: dict[str, int]
    source_count: int

    def to_json(self) -> dict[str, Any]:
        return {
            "chunks": [chunk.to_json() for chunk in self.chunks],
            "warnings": self.warnings,
            "redaction_counts": self.redaction_counts,
            "source_count": self.source_count,
            "character_count": sum(len(chunk.text) for chunk in self.chunks),
        }


def load_evidence_json(path: str) -> EvidenceBundle:
    raw = Path(path).read_text(encoding="utf-8")
    payload = json.loads(raw)
    rows = payload.get("chunks") if isinstance(payload, dict) else payload
    if not isinstance(rows, list):
        raise ValueError("evidence JSON must be an array or an object with chunks")
    chunks: list[EvidenceChunk] = []
    for index, row in enumerate(rows, 1):
        if not isinstance(row, dict) or not str(row.get("text") or "").strip():
            continue
        chunks.append(
            EvidenceChunk(
                id=str(row.get("id") or f"E{index:03d}"),
                source_name=str(row.get("source_name") or "evidence"),
                text=str(row["text"]),
                kind=str(row.get("kind") or "text"),
                page=_optional_int(row.get("page")),
                row=_optional_int(row.get("row")),
                source_url=_optional_string(row.get("source_url")),
                local_path=_optional_string(row.get("local_path")),
                extraction_warning=_optional_string(row.get("extraction_warning")),
            )
        )
    return EvidenceBundle(
        chunks=chunks,
        warnings=[str(item) for item in payload.get("warnings", [])]
        if isinstance(payload, dict)
        else [],
        redaction_counts=dict(payload.get("redaction_counts") or {})
        if isinstance(payload, dict)
        else {},
        source_count=int(payload.get("source_count") or len(chunks))
        if isinstance(payload, dict)
        else len(chunks),
    )


def evidence_from_materials(
    manifest: dict[str, Any],
    *,
    chunk_chars: int = DEFAULT_CHUNK_CHARS,
    max_evidence_chars: int = DEFAULT_MAX_EVIDENCE_CHARS,
    redact: bool = True,
) -> EvidenceBundle:
    raw_chunks: list[dict[str, Any]] = []
    warnings: list[str] = []
    source_count = 0

    detail = manifest.get("detail") or {}
    if isinstance(detail, dict):
        detail_text = str(detail.get("text_preview") or "").strip()
        if detail_text:
            source_count += 1
            raw_chunks.extend(
                _split_source_text(
                    detail_text,
                    source_name="공지 본문",
                    kind="html",
                    source_url=_optional_string(detail.get("url")),
                    local_path=_optional_string(detail.get("local_path")),
                    chunk_chars=chunk_chars,
                )
            )

    attachments = manifest.get("attachments") or []
    for index, attachment in enumerate(attachments):
        if not isinstance(attachment, dict):
            continue
        local_path = attachment.get("local_path")
        if not local_path or attachment.get("fetch_status") not in {None, "ok"}:
            continue
        path = Path(str(local_path))
        if not path.exists():
            warnings.append(f"attachment path does not exist: {path}")
            continue
        source_count += 1
        source_name = str(attachment.get("name") or path.name or f"attachment-{index}")
        try:
            extracted, extracted_warnings = extract_path(
                path,
                source_name=source_name,
                source_url=_optional_string(attachment.get("url")),
                chunk_chars=chunk_chars,
            )
            raw_chunks.extend(extracted)
            warnings.extend(extracted_warnings)
        except Exception as error:  # noqa: BLE001 - retain other usable evidence.
            warnings.append(f"failed to extract {source_name}: {error}")

    chunks: list[EvidenceChunk] = []
    redaction_counts = {"email": 0, "phone": 0, "resident_id": 0, "student_id": 0}
    used_chars = 0
    for raw in raw_chunks:
        text = str(raw.get("text") or "").strip()
        if not text:
            continue
        if redact:
            text, counts = redact_sensitive_text(text)
            for key, count in counts.items():
                redaction_counts[key] += count
        if used_chars + len(text) > max_evidence_chars:
            remaining = max_evidence_chars - used_chars
            if remaining < 200:
                warnings.append("evidence character limit reached")
                break
            text = text[:remaining].rstrip()
            warnings.append("last evidence chunk was truncated at the character limit")
        chunks.append(
            EvidenceChunk(
                id=f"E{len(chunks) + 1:03d}",
                source_name=str(raw.get("source_name") or "evidence"),
                text=text,
                kind=str(raw.get("kind") or "text"),
                page=_optional_int(raw.get("page")),
                row=_optional_int(raw.get("row")),
                source_url=_optional_string(raw.get("source_url")),
                local_path=_optional_string(raw.get("local_path")),
                extraction_warning=_optional_string(raw.get("extraction_warning")),
            )
        )
        used_chars += len(text)
        if used_chars >= max_evidence_chars:
            break

    return EvidenceBundle(
        chunks=chunks,
        warnings=warnings,
        redaction_counts=redaction_counts,
        source_count=source_count,
    )


def extract_path(
    path: Path,
    *,
    source_name: str,
    source_url: str | None,
    chunk_chars: int,
    zip_depth: int = 0,
) -> tuple[list[dict[str, Any]], list[str]]:
    suffix = path.suffix.casefold()
    if suffix in {".txt", ".md", ".log"}:
        text = read_text_file(path)
        return (
            _split_source_text(
                text,
                source_name=source_name,
                kind="text",
                source_url=source_url,
                local_path=str(path),
                chunk_chars=chunk_chars,
            ),
            [],
        )
    if suffix in {".html", ".htm"}:
        text = extract_html_text(read_text_file(path))
        return (
            _split_source_text(
                text,
                source_name=source_name,
                kind="html",
                source_url=source_url,
                local_path=str(path),
                chunk_chars=chunk_chars,
            ),
            [],
        )
    if suffix == ".pdf":
        return extract_pdf(path, source_name=source_name, source_url=source_url, chunk_chars=chunk_chars)
    if suffix in {".xlsx", ".xlsm"}:
        return extract_xlsx(path, source_name=source_name, source_url=source_url)
    if suffix == ".csv":
        return extract_csv(path, source_name=source_name, source_url=source_url)
    if suffix == ".hwp":
        return extract_hwp(path, source_name=source_name, source_url=source_url, chunk_chars=chunk_chars)
    if suffix == ".zip":
        return extract_zip(
            path,
            source_name=source_name,
            source_url=source_url,
            chunk_chars=chunk_chars,
            zip_depth=zip_depth,
        )
    if suffix in {".jpg", ".jpeg", ".png", ".webp"}:
        ocr_text, ocr_warning = try_local_ocr_image(path)
        if ocr_text:
            return (
                _split_source_text(
                    ocr_text,
                    source_name=source_name,
                    kind="image_ocr",
                    source_url=source_url,
                    local_path=str(path),
                    chunk_chars=chunk_chars,
                ),
                [ocr_warning] if ocr_warning else [],
            )
        return (
            [
                {
                    "source_name": source_name,
                    "kind": "image_visual",
                    "text": VISUAL_PENDING_TEXT,
                    "source_url": source_url,
                    "local_path": str(path),
                }
            ],
            [f"image queued for multimodal transcription: {source_name}"],
        )
    return [], [f"unsupported attachment format {suffix or '(none)'}: {source_name}"]


def extract_pdf(
    path: Path,
    *,
    source_name: str,
    source_url: str | None,
    chunk_chars: int,
) -> tuple[list[dict[str, Any]], list[str]]:
    try:
        from pypdf import PdfReader
    except ImportError as error:
        raise RuntimeError("PDF extraction requires pypdf") from error
    reader = PdfReader(path)
    chunks: list[dict[str, Any]] = []
    warnings: list[str] = []
    nonempty_pages = 0
    for page_number, page in enumerate(reader.pages, 1):
        text = (page.extract_text() or "").strip()
        if not text:
            ocr_text, ocr_warning = try_local_ocr_pdf_page(path, page_number)
            if ocr_text:
                chunks.extend(
                    _split_source_text(
                        ocr_text,
                        source_name=source_name,
                        kind="pdf_ocr_page",
                        page=page_number,
                        source_url=source_url,
                        local_path=str(path),
                        chunk_chars=chunk_chars,
                    )
                )
                nonempty_pages += 1
                if ocr_warning:
                    warnings.append(ocr_warning)
                continue
            chunks.append(
                {
                    "source_name": source_name,
                    "kind": "pdf_visual_page",
                    "page": page_number,
                    "text": VISUAL_PENDING_TEXT,
                    "source_url": source_url,
                    "local_path": str(path),
                }
            )
            continue
        nonempty_pages += 1
        chunks.extend(
            _split_source_text(
                text,
                source_name=source_name,
                kind="pdf_page",
                page=page_number,
                source_url=source_url,
                local_path=str(path),
                chunk_chars=chunk_chars,
            )
        )
    if reader.pages and nonempty_pages == 0:
        warnings.append(f"PDF has no text layer; all pages queued for multimodal transcription: {source_name}")
    elif nonempty_pages < len(reader.pages):
        warnings.append(
            f"PDF has {len(reader.pages) - nonempty_pages} page(s) queued for multimodal transcription: {source_name}"
        )
    return chunks, warnings


def try_local_ocr_pdf_page(path: Path, page_number: int) -> tuple[str | None, str | None]:
    if not tesseract_executable():
        return None, None
    try:
        import fitz
    except ImportError:
        return None, "local PDF OCR skipped because PyMuPDF is unavailable"

    try:
        with fitz.open(path) as document, tempfile.TemporaryDirectory() as temp_dir:
            page = document.load_page(page_number - 1)
            pixmap = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
            image_path = Path(temp_dir) / f"page-{page_number}.png"
            pixmap.save(image_path)
            text, warning = try_local_ocr_image(image_path)
    except Exception as error:  # noqa: BLE001 - VLM fallback remains available.
        return None, f"local OCR failed for PDF page {page_number}: {error}"
    if not text:
        return None, warning
    return text, f"local OCR used for PDF page {page_number}"


def try_local_ocr_image(path: Path) -> tuple[str | None, str | None]:
    executable = tesseract_executable()
    if not executable:
        return None, None
    errors: list[str] = []
    for language in ("kor+eng", "eng"):
        try:
            completed = subprocess.run(
                [
                    executable,
                    str(path),
                    "stdout",
                    "-l",
                    language,
                    "--psm",
                    "6",
                ],
                capture_output=True,
                check=False,
                timeout=60,
            )
        except Exception as error:  # noqa: BLE001 - VLM fallback remains available.
            errors.append(str(error))
            continue
        if completed.returncode != 0:
            errors.append(completed.stderr.decode("utf-8", errors="replace").strip())
            continue
        text = completed.stdout.decode("utf-8", errors="replace").strip()
        if len(normalize_text(text)) >= MIN_LOCAL_OCR_CHARS:
            return text, f"local OCR used ({language}): {path.name}"
    warning = "; ".join(error for error in errors if error)
    return None, f"local OCR produced no usable text: {warning}" if warning else None


def tesseract_executable() -> str | None:
    return shutil.which("tesseract") or shutil.which("tesseract.exe")


def extract_xlsx(
    path: Path,
    *,
    source_name: str,
    source_url: str | None,
) -> tuple[list[dict[str, Any]], list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise RuntimeError("XLSX extraction requires openpyxl") from error
    workbook = load_workbook(path, read_only=True, data_only=True)
    chunks: list[dict[str, Any]] = []
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        header_index = _detect_header_row(rows[:20])
        headers = [_cell_text(value) or f"column_{index + 1}" for index, value in enumerate(rows[header_index])]
        for row_number, values in enumerate(rows[header_index + 1 :], header_index + 2):
            pairs = [
                f"{headers[index]}={_cell_text(value)}"
                for index, value in enumerate(values)
                if index < len(headers) and _cell_text(value)
            ]
            if not pairs:
                continue
            chunks.append(
                {
                    "source_name": f"{source_name} / {sheet.title}",
                    "kind": "xlsx_row",
                    "row": row_number,
                    "text": " | ".join(pairs),
                    "source_url": source_url,
                    "local_path": str(path),
                }
            )
    workbook.close()
    return chunks, []


def extract_csv(
    path: Path,
    *,
    source_name: str,
    source_url: str | None,
) -> tuple[list[dict[str, Any]], list[str]]:
    text = read_text_file(path)
    rows = list(csv.reader(io.StringIO(text)))
    if not rows:
        return [], []
    headers = [value.strip() or f"column_{index + 1}" for index, value in enumerate(rows[0])]
    chunks = []
    for row_number, values in enumerate(rows[1:], 2):
        pairs = [
            f"{headers[index]}={value.strip()}"
            for index, value in enumerate(values)
            if index < len(headers) and value.strip()
        ]
        if pairs:
            chunks.append(
                {
                    "source_name": source_name,
                    "kind": "csv_row",
                    "row": row_number,
                    "text": " | ".join(pairs),
                    "source_url": source_url,
                    "local_path": str(path),
                }
            )
    return chunks, []


def extract_hwp(
    path: Path,
    *,
    source_name: str,
    source_url: str | None,
    chunk_chars: int,
) -> tuple[list[dict[str, Any]], list[str]]:
    executable = shutil.which("hwp5txt") or shutil.which("hwp5txt.exe")
    if not executable:
        return [], [f"HWP extraction requires hwp5txt: {source_name}"]
    completed = subprocess.run(
        [executable, str(path)],
        capture_output=True,
        check=False,
        timeout=60,
    )
    if completed.returncode != 0:
        return [], [f"hwp5txt failed for {source_name}: exit {completed.returncode}"]
    text = completed.stdout.decode("utf-8", errors="replace")
    return (
        _split_source_text(
            text,
            source_name=source_name,
            kind="hwp_text",
            source_url=source_url,
            local_path=str(path),
            chunk_chars=chunk_chars,
        ),
        [],
    )


def extract_zip(
    path: Path,
    *,
    source_name: str,
    source_url: str | None,
    chunk_chars: int,
    zip_depth: int,
) -> tuple[list[dict[str, Any]], list[str]]:
    if zip_depth >= 2:
        return [], [f"nested ZIP depth limit reached: {source_name}"]
    chunks: list[dict[str, Any]] = []
    warnings: list[str] = []
    total_bytes = 0
    with zipfile.ZipFile(path) as archive, tempfile.TemporaryDirectory() as temp_dir:
        root = Path(temp_dir).resolve()
        for info in archive.infolist():
            if info.is_dir():
                continue
            total_bytes += info.file_size
            if info.file_size > DEFAULT_MAX_ZIP_ENTRY_BYTES:
                warnings.append(f"ZIP entry too large: {info.filename}")
                continue
            if total_bytes > DEFAULT_MAX_ZIP_TOTAL_BYTES:
                warnings.append(f"ZIP expanded size limit reached: {source_name}")
                break
            target = (root / info.filename).resolve()
            if root not in target.parents:
                warnings.append(f"unsafe ZIP path skipped: {info.filename}")
                continue
            target.parent.mkdir(parents=True, exist_ok=True)
            target.write_bytes(archive.read(info))
            child_name = f"{source_name} / {info.filename}"
            child_chunks, child_warnings = extract_path(
                target,
                source_name=child_name,
                source_url=source_url,
                chunk_chars=chunk_chars,
                zip_depth=zip_depth + 1,
            )
            chunks.extend(child_chunks)
            warnings.extend(child_warnings)
    return chunks, warnings


def redact_sensitive_text(value: str) -> tuple[str, dict[str, int]]:
    patterns = {
        "email": re.compile(r"\b[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}\b", re.IGNORECASE),
        "resident_id": re.compile(r"\b\d{6}\s*-\s*[1-4]\d{6}\b"),
        "phone": re.compile(r"(?<!\d)(?:\+?82[-\s]?)?(?:0\d{1,2})[-\s]?\d{3,4}[-\s]?\d{4}(?!\d)"),
        # PNU student numbers are normally 9-10 digits. Avoid broad patterns that
        # redact dates such as 20260719 and destroy deadline evidence.
        "student_id": re.compile(r"(?<!\d)\d{9,10}(?!\d)"),
    }
    replacements = {
        "email": "[EMAIL_REDACTED]",
        "resident_id": "[RESIDENT_ID_REDACTED]",
        "phone": "[PHONE_REDACTED]",
        "student_id": "[STUDENT_ID_REDACTED]",
    }
    counts: dict[str, int] = {}
    text = value
    for key in ("email", "resident_id", "phone", "student_id"):
        text, count = patterns[key].subn(replacements[key], text)
        counts[key] = count
    return text, counts


def lexical_rank(
    request: str,
    chunks: list[EvidenceChunk],
    *,
    limit: int,
) -> list[tuple[EvidenceChunk, float]]:
    query = normalize_text(request)
    query_tokens = search_tokens(query)
    query_numbers = set(re.findall(r"\d+", query))
    ranked: list[tuple[EvidenceChunk, float]] = []
    for chunk in chunks:
        text = normalize_text(chunk.text)
        tokens = search_tokens(text)
        overlap = query_tokens.intersection(tokens)
        score = float(len(overlap) * 3)
        if query and query in text:
            score += 12
        for number in query_numbers:
            if re.search(rf"(?<!\d){re.escape(number)}(?!\d)", text):
                score += 4
        compact_query_terms = [term.replace(" ", "") for term in query_tokens if len(term) >= 3]
        compact_text = text.replace(" ", "")
        score += sum(2 for term in compact_query_terms if term in compact_text)
        if chunk.kind in {"xlsx_row", "csv_row"} and overlap:
            score += 2
        ranked.append((chunk, score))
    ranked.sort(key=lambda item: (-item[1], item[0].id))
    return ranked[: max(1, limit)]


def search_tokens(value: str) -> set[str]:
    tokens = set(re.findall(r"[가-힣A-Za-z]+|\d+", value))
    return {token for token in tokens if len(token) >= 2 and token not in STOP_TERMS}


def cosine_similarity(left: list[float], right: list[float]) -> float:
    if len(left) != len(right) or not left:
        return 0.0
    dot = sum(a * b for a, b in zip(left, right))
    left_norm = math.sqrt(sum(value * value for value in left))
    right_norm = math.sqrt(sum(value * value for value in right))
    if left_norm == 0 or right_norm == 0:
        return 0.0
    return dot / (left_norm * right_norm)


def read_text_file(path: Path) -> str:
    data = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "cp949", "euc-kr"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def extract_html_text(value: str) -> str:
    parser = _VisibleTextParser()
    parser.feed(value)
    return re.sub(r"\s+", " ", " ".join(parser.parts)).strip()


def normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", unicodedata.normalize("NFKC", value).casefold()).strip()


def _split_source_text(
    text: str,
    *,
    source_name: str,
    kind: str,
    source_url: str | None,
    local_path: str | None,
    chunk_chars: int,
    page: int | None = None,
) -> list[dict[str, Any]]:
    normalized = re.sub(r"\r\n?", "\n", text).strip()
    if not normalized:
        return []
    paragraphs = [part.strip() for part in re.split(r"\n\s*\n|(?<=[.!?。])\s+", normalized) if part.strip()]
    chunks: list[str] = []
    current = ""
    for paragraph in paragraphs:
        if len(paragraph) > chunk_chars:
            if current:
                chunks.append(current)
                current = ""
            chunks.extend(
                paragraph[index : index + chunk_chars]
                for index in range(0, len(paragraph), chunk_chars)
            )
            continue
        candidate = f"{current}\n{paragraph}".strip() if current else paragraph
        if len(candidate) > chunk_chars and current:
            chunks.append(current)
            current = paragraph
        else:
            current = candidate
    if current:
        chunks.append(current)
    return [
        {
            "source_name": source_name,
            "kind": kind,
            "page": page,
            "text": chunk,
            "source_url": source_url,
            "local_path": local_path,
        }
        for chunk in chunks
    ]


def _detect_header_row(rows: list[tuple[Any, ...]]) -> int:
    if not rows:
        return 0
    scores = []
    for index, row in enumerate(rows):
        nonempty = sum(1 for value in row if _cell_text(value))
        text_cells = sum(1 for value in row if isinstance(value, str) and value.strip())
        scores.append((nonempty + text_cells, -index, index))
    return max(scores)[2]


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _optional_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    return int(value)


def _optional_string(value: Any) -> str | None:
    if value is None or value == "":
        return None
    return str(value)


class _VisibleTextParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.parts: list[str] = []
        self.ignored_depth = 0

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.casefold() in {"script", "style", "noscript"}:
            self.ignored_depth += 1

    def handle_endtag(self, tag: str) -> None:
        if tag.casefold() in {"script", "style", "noscript"} and self.ignored_depth:
            self.ignored_depth -= 1

    def handle_data(self, data: str) -> None:
        if not self.ignored_depth and data.strip():
            self.parts.append(data.strip())
